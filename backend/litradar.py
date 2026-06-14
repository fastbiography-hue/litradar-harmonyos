#!/usr/bin/env python3
"""
LitRadar v3.1 — 融合版生物医学文献追踪系统
=============================================
SQLite 做脑 · Excel 做脸 · 定时追踪做心跳

v3.1 更新:
  - 安全: API Key / 邮箱密码 支持环境变量
  - bioRxiv: paperscraper 服务端关键词搜索
  - Europe PMC: MeSH Terms 解析
  - upsert: 插入前预检查，rowcount 不再误判
  - PubMed: MeSH 词自动扩展提升召回率
  - 去重: 加入 Title 模糊匹配 (difflib)
  - BibTeX 导出
  - 邮件: HTML+纯文本双part, Excel附件, 统计Banner, 高相关专区
  - search 命令: --send-email 参数
"""

import argparse
import difflib
import hashlib
import json
import os
import re
import smtplib
import sqlite3
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor
from contextlib import contextmanager
from datetime import datetime, timedelta
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import Optional

import pandas as pd
import requests
import yaml

try:
    from loguru import logger
except ImportError:
    import logging
    logger = logging.getLogger("litradar")
    logger.setLevel(logging.INFO)
    _h = logging.StreamHandler()
    _h.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
    logger.addHandler(_h)


# ===========================================================================
# 配置
# ===========================================================================


def load_config(config_path: str = "config.yaml") -> dict:
    if not os.path.exists(config_path):
        logger.error(f"配置文件不存在: {config_path}")
        sys.exit(1)
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f) or {}

    # 安全: API Key 优先读环境变量
    cfg["ncbi_api_key"] = cfg.get("ncbi_api_key") or os.environ.get("NCBI_API_KEY", "")
    cfg["ncbi_email"] = cfg.get("ncbi_email") or os.environ.get("NCBI_EMAIL", "user@example.com")

    # 邮件密码优先读环境变量
    ec = cfg.get("notifications", {}).get("email", {})
    if not ec.get("password"):
        ec["password"] = os.environ.get("LITRADAR_EMAIL_PASS", "")
    return cfg


# ===========================================================================
# SQLite 存储
# ===========================================================================


class Storage:
    def __init__(self, db_path: str = "data/litradar.db"):
        self.db_path = db_path
        os.makedirs(os.path.dirname(db_path), exist_ok=True)
        self._init_db()

    @contextmanager
    def _conn(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        try:
            yield conn
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def _init_db(self):
        with self._conn() as c:
            c.executescript("""
                CREATE TABLE IF NOT EXISTS papers (
                    id              TEXT PRIMARY KEY,
                    Title           TEXT,
                    "First Author"  TEXT,
                    "All Authors"   TEXT,
                    Journal         TEXT,
                    Year            TEXT,
                    "Pub Date"      TEXT,
                    Abstract        TEXT,
                    "MeSH Terms"    TEXT,
                    Keywords        TEXT,
                    DOI             TEXT,
                    PMID            TEXT,
                    PMCID           TEXT,
                    "Is OA"         TEXT DEFAULT 'No',
                    Source          TEXT,
                    Citations       INTEGER DEFAULT 0,
                    "Query Tag"     TEXT,
                    "Relevance Score" REAL DEFAULT 0.0,
                    pdf_path        TEXT,
                    pdf_md5         TEXT,
                    created_at      TEXT,
                    updated_at      TEXT
                );
                CREATE INDEX IF NOT EXISTS idx_doi ON papers(DOI);
                CREATE INDEX IF NOT EXISTS idx_pmid ON papers(PMID);
                CREATE INDEX IF NOT EXISTS idx_source ON papers(Source);
                CREATE INDEX IF NOT EXISTS idx_query_tag ON papers("Query Tag");
                CREATE INDEX IF NOT EXISTS idx_score ON papers("Relevance Score");

                CREATE TABLE IF NOT EXISTS track_runs (
                    id          INTEGER PRIMARY KEY AUTOINCREMENT,
                    query_name  TEXT,
                    query_str   TEXT,
                    run_at      TEXT,
                    papers_found INTEGER DEFAULT 0,
                    papers_new  INTEGER DEFAULT 0
                );
            """)

    def upsert(self, paper: dict) -> bool:
        """插入前先检查是否存在，可靠判断是否新增。"""
        now = datetime.now().isoformat()
        paper.setdefault("created_at", now)
        paper["updated_at"] = now

        fields = [
            "id", "Title", "First Author", "All Authors", "Journal", "Year",
            "Pub Date", "Abstract", "MeSH Terms", "Keywords", "DOI", "PMID",
            "PMCID", "Is OA", "Source", "Citations", "Query Tag",
            "Relevance Score", "pdf_path", "pdf_md5", "created_at", "updated_at",
        ]
        for f in fields:
            paper.setdefault(f, "" if f not in ("Citations", "Relevance Score") else 0)
        for f in ("DOI", "PMID", "PMCID"):
            if paper.get(f) == "":
                paper[f] = None

        # 插入前检查
        is_new = not self.exists(paper.get("DOI", ""), paper.get("PMID", ""))

        qfields = [f if f == "id" else f'"{f}"' for f in fields]
        pl = ", ".join(["?" for _ in fields])
        up = ", ".join([f'"{f}"=excluded."{f}"' for f in fields if f != "id"])
        values = [paper.get(f) for f in fields]
        try:
            with self._conn() as conn:
                conn.execute(
                    f'INSERT INTO papers ({", ".join(qfields)}) VALUES ({pl}) '
                    f"ON CONFLICT(id) DO UPDATE SET {up}",
                    values,
                )
                return is_new
        except Exception as e:
            logger.debug(f"upsert 失败 [{paper.get('id','')}]: {e}")
            return False

    def exists(self, doi: str = "", pmid: str = "") -> bool:
        with self._conn() as c:
            if doi:
                r = c.execute("SELECT 1 FROM papers WHERE DOI=?", (doi,)).fetchone()
                if r: return True
            if pmid:
                r = c.execute("SELECT 1 FROM papers WHERE PMID=?", (pmid,)).fetchone()
                if r: return True
        return False

    def get_all(self, limit: int = 10000) -> list[dict]:
        with self._conn() as c:
            rows = c.execute(
                'SELECT * FROM papers ORDER BY "Relevance Score" DESC, "Pub Date" DESC LIMIT ?',
                (limit,),
            ).fetchall()
        return [dict(r) for r in rows]

    def get_undownloaded(self) -> list[dict]:
        with self._conn() as c:
            rows = c.execute(
                'SELECT * FROM papers WHERE "Is OA"=\'Yes\' AND (pdf_path IS NULL OR pdf_path=\'\')'
            ).fetchall()
        return [dict(r) for r in rows]

    def mark_downloaded(self, paper_id: str, pdf_path: str):
        md5 = self._file_md5(pdf_path)
        with self._conn() as c:
            c.execute(
                "UPDATE papers SET pdf_path=?, pdf_md5=?, updated_at=? WHERE id=?",
                (pdf_path, md5, datetime.now().isoformat(), paper_id),
            )

    def get_stats(self) -> dict:
        with self._conn() as c:
            total = c.execute("SELECT COUNT(*) FROM papers").fetchone()[0]
            oa = c.execute('SELECT COUNT(*) FROM papers WHERE "Is OA"=\'Yes\'').fetchone()[0]
            dl = c.execute(
                "SELECT COUNT(*) FROM papers WHERE pdf_path IS NOT NULL AND pdf_path!=''"
            ).fetchone()[0]
            high = c.execute(
                'SELECT COUNT(*) FROM papers WHERE "Relevance Score">=7'
            ).fetchone()[0]
        return {"total": total, "oa": oa, "downloaded": dl, "high_relevance": high}

    def log_track_run(self, query_name: str, query_str: str, found: int, new: int):
        with self._conn() as c:
            c.execute(
                "INSERT INTO track_runs (query_name, query_str, run_at, papers_found, papers_new) "
                "VALUES (?, ?, ?, ?, ?)",
                (query_name, query_str, datetime.now().isoformat(), found, new),
            )

    def get_last_run(self, query_name: str) -> Optional[str]:
        with self._conn() as c:
            r = c.execute(
                "SELECT run_at FROM track_runs WHERE query_name=? ORDER BY run_at DESC LIMIT 1",
                (query_name,),
            ).fetchone()
        return r["run_at"] if r else None

    def get_track_history(self) -> list[dict]:
        with self._conn() as c:
            rows = c.execute(
                "SELECT * FROM track_runs ORDER BY run_at DESC LIMIT 50"
            ).fetchall()
        return [dict(r) for r in rows]

    @staticmethod
    def _file_md5(path: str) -> str:
        h = hashlib.md5()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                h.update(chunk)
        return h.hexdigest()


# ===========================================================================
# 数据源检索
# ===========================================================================

_rate_ncbi_last = 0.0

def _rate_limit_ncbi(api_key: str):
    global _rate_ncbi_last
    rate = 10.0 if api_key else 3.0
    now = time.monotonic()
    elapsed = now - _rate_ncbi_last
    if elapsed < 1.0 / rate:
        time.sleep(1.0 / rate - elapsed)
    _rate_ncbi_last = time.monotonic()


# -- MeSH 自动扩展 --
MESH_EXPANSIONS = {
    "Alzheimer": '"Alzheimer Disease"[MeSH]',
    "Parkinson": '"Parkinson Disease"[MeSH]',
    "diabetes": '"Diabetes Mellitus"[MeSH]',
    "obesity": '"Obesity"[MeSH]',
    "hypertension": '"Hypertension"[MeSH]',
    "cancer": '"Neoplasms"[MeSH]',
    "stroke": '"Stroke"[MeSH]',
    "schizophrenia": '"Schizophrenia"[MeSH]',
    "depression": '"Depressive Disorder"[MeSH]',
    "APOE": '"Apolipoprotein E4"[MeSH] OR "Apolipoproteins E"[MeSH]',
}


def _expand_query_with_mesh(query: str) -> str:
    """自动为查询中的关键词添加 MeSH 词扩展，提升 PubMed 召回率。"""
    expanded = query
    for term, mesh in MESH_EXPANSIONS.items():
        if term.lower() in query.lower() and mesh not in expanded:
            expanded += f" OR {mesh}"
    return expanded


# -- PubMed --

def fetch_pubmed(query: str, config: dict, date_from: str) -> list[dict]:
    try:
        from Bio import Entrez
    except ImportError:
        logger.warning("BioPython 未安装，跳过 PubMed")
        return []

    ncbi_email = config.get("ncbi_email", "user@example.com")
    ncbi_api_key = config.get("ncbi_api_key", "")
    max_results = config.get("max_results_per_source", 200)
    Entrez.email = ncbi_email
    if ncbi_api_key:
        Entrez.api_key = ncbi_api_key

    # MeSH 自动扩展
    full_query = _expand_query_with_mesh(query)
    if date_from:
        full_query += f' AND ("{date_from}"[Date - Publication] : "3000"[Date - Publication])'

    logger.info(f"  [PubMed] {full_query[:120]}...")

    _rate_limit_ncbi(ncbi_api_key)
    try:
        handle = Entrez.esearch(db="pubmed", term=full_query, retmax=max_results,
                                sort="relevance", usehistory="y")
        result = Entrez.read(handle); handle.close()
    except Exception as e:
        logger.error(f"  [PubMed] esearch 失败: {e}")
        return []

    ids = result.get("IdList", [])
    if not ids:
        logger.info("  [PubMed] 无结果")
        return []
    logger.info(f"  [PubMed] 找到 {len(ids)} 条")

    papers = []
    for i in range(0, len(ids), 200):
        batch = ids[i:i + 200]
        _rate_limit_ncbi(ncbi_api_key)
        try:
            handle = Entrez.efetch(db="pubmed", id=",".join(batch), rettype="xml", retmode="xml")
            articles = Entrez.read(handle); handle.close()
            for art in articles.get("PubmedArticle", []):
                p = _parse_pubmed(art)
                if p: papers.append(p)
        except Exception as e:
            logger.error(f"  [PubMed] efetch 失败: {e}")
        if i + 200 < len(ids):
            time.sleep(0.35)

    logger.info(f"  [PubMed] → {len(papers)} 篇")
    return papers


def _parse_pubmed(article: dict) -> Optional[dict]:
    try:
        med = article.get("MedlineCitation", {})
        art = med.get("Article", {})
        pub = article.get("PubmedData", {})
        pmid = str(med.get("PMID", ""))

        title = str(art.get("ArticleTitle", ""))
        authors = []
        for a in art.get("AuthorList", []):
            l, f = a.get("LastName", ""), a.get("ForeName", "")
            if l or f: authors.append(f"{l} {f}".strip())
        fa = authors[0] if authors else ""

        abs_parts = art.get("Abstract", {}).get("AbstractText", [])
        if isinstance(abs_parts, list):
            abstract = " ".join(
                str(p) if isinstance(p, str) else p.get("#text", str(p)) for p in abs_parts
            )
        else:
            abstract = str(abs_parts)

        jn = str(art.get("Journal", {}).get("Title", ""))
        di = art.get("Journal", {}).get("JournalIssue", {}).get("PubDate", {})
        yr = str(di.get("Year", ""))
        mo = str(di.get("Month", "01")).zfill(2)
        dy = str(di.get("Day", "01")).zfill(2)
        pd = f"{yr}-{mo}-{dy}" if yr else ""

        doi = ""
        for e in art.get("ELocationID", []):
            if e.attributes.get("EIdType") == "doi":
                doi = str(e)

        mesh = []
        for mh in med.get("MeshHeadingList", []):
            d = mh.get("DescriptorName", {})
            if d: mesh.append(str(d))

        pmcid, oa = "", "No"
        for pid in pub.get("ArticleIdList", []):
            if pid.attributes.get("IdType") == "pmc":
                pmcid = str(pid); oa = "Yes"

        paper_id = f"pmid:{pmid}" if pmid else f"pubmed:{hashlib.md5(title.encode()).hexdigest()[:12]}"
        return {
            "id": paper_id, "Title": title, "First Author": fa,
            "All Authors": "; ".join(authors), "Journal": jn,
            "Year": yr, "Pub Date": pd, "Abstract": abstract,
            "MeSH Terms": "; ".join(mesh), "Keywords": "",
            "DOI": doi, "PMID": pmid, "PMCID": pmcid,
            "Is OA": oa, "Source": "PubMed", "Citations": 0,
            "Query Tag": "", "Relevance Score": 0.0,
        }
    except Exception:
        return None


# -- Europe PMC --

def fetch_europe_pmc(query: str, config: dict, date_from: str) -> list[dict]:
    max_results = config.get("max_results_per_source", 200)
    email = config.get("ncbi_email", "")
    full_q = query
    if date_from:
        full_q += f" AND FIRST_PDATE:[{date_from} TO 3000-12-31]"

    logger.info(f"  [Europe PMC] {full_q[:100]}...")

    papers = []
    cursor = "*"
    while len(papers) < max_results:
        resp = requests.get(
            "https://www.ebi.ac.uk/europepmc/webservices/rest/search",
            params={"query": full_q, "resultType": "core", "pageSize": min(max_results, 200),
                    "format": "json", "cursorMark": cursor},
            headers={"User-Agent": f"LitRadar/3.1 (mailto:{email})"},
            timeout=30,
        )
        if resp.status_code != 200: break
        data = resp.json()
        results = data.get("resultList", {})
        items = results.get("result", [])

        for r in items:
            has_oa = r.get("isOpenAccess", "") == "Y"

            # 解析 MeSH Terms
            mesh = []
            mesh_list = r.get("meshHeadingList", {})
            if isinstance(mesh_list, dict):
                for mh in mesh_list.get("meshHeading", []):
                    dn = mh.get("descriptorName", "")
                    if dn: mesh.append(dn)

            papers.append({
                "id": f"pmid:{r['pmid']}" if r.get("pmid") else f"epmc:{hashlib.md5((r.get('title') or '').encode()).hexdigest()[:12]}",
                "Title": r.get("title", "") or "",
                "First Author": ((r.get("authorString") or "").split(",")[0].strip()),
                "All Authors": (r.get("authorString") or "").replace(",", "; "),
                "Journal": r.get("journalTitle", "") or "",
                "Year": str(r.get("pubYear", "")),
                "Pub Date": r.get("firstPublicationDate", "") or "",
                "Abstract": r.get("abstractText", "") or "",
                "MeSH Terms": "; ".join(mesh),
                "Keywords": "",
                "DOI": r.get("doi", "") or "",
                "PMID": r.get("pmid", "") or "",
                "PMCID": r.get("pmcid", "") or "",
                "Is OA": "Yes" if has_oa else "No",
                "Source": "Europe PMC", "Citations": int(r.get("citedByCount", 0) or 0),
                "Query Tag": "", "Relevance Score": 0.0,
            })

        nc = results.get("nextCursorMark", "")
        if not nc or nc == cursor or not items: break
        cursor = nc
        time.sleep(0.2)

    logger.info(f"  [Europe PMC] → {len(papers)} 篇")
    return papers


# -- bioRxiv / medRxiv --

def fetch_biorxiv(query: str, config: dict, date_from: str) -> list[dict]:
    return _fetch_rxiv("biorxiv", query, config, date_from)

def fetch_medrxiv(query: str, config: dict, date_from: str) -> list[dict]:
    return _fetch_rxiv("medrxiv", query, config, date_from)


def _fetch_rxiv(server: str, query: str, config: dict, date_from: str) -> list[dict]:
    max_results = config.get("max_results_per_source", 200)
    email = config.get("ncbi_email", "")
    logger.info(f"  [{server}] {query[:100]}...")

    papers = []

    # 方案 A: paperscraper 服务端关键词搜索
    try:
        from paperscraper.xrxiv.xrxiv_query import XRXivQuery
        q = XRXivQuery(server)
        # 将布尔查询转为关键词列表
        keywords = _extract_keywords_from_query(query)
        results = q.text_search(" ".join(keywords) if keywords else query, limit=max_results)
        if isinstance(results, list):
            for r in results:
                if not isinstance(r, dict): continue
                doi = r.get("doi", ""); title = r.get("title", "")
                authors = r.get("authors", "")
                if isinstance(authors, list):
                    au_str = "; ".join(authors); fa = authors[0] if authors else ""
                else:
                    au_str = str(authors); fa = au_str.split(";")[0].strip()
                papers.append({
                    "id": f"doi:{doi}" if doi else f"{server}:{hashlib.md5(title.encode()).hexdigest()[:12]}",
                    "Title": title, "First Author": fa, "All Authors": au_str,
                    "Journal": f"{server} (preprint)",
                    "Year": str(r.get("date", ""))[:4], "Pub Date": str(r.get("date", "")),
                    "Abstract": r.get("abstract", ""), "MeSH Terms": "", "Keywords": "",
                    "DOI": doi, "PMID": "", "PMCID": "", "Is OA": "Yes",
                    "Source": server, "Citations": 0, "Query Tag": "", "Relevance Score": 0.0,
                })
        if papers:
            logger.info(f"  [{server}] (paperscraper) → {len(papers)} 篇")
            return papers
    except ImportError:
        pass
    except Exception as e:
        logger.debug(f"  [{server}] paperscraper: {e}")

    # 方案 B: REST API + 客户端过滤
    url = f"https://api.biorxiv.org/details/{server}/2020-01-01/3000-12-31/0/{min(max_results, 200)}"
    try:
        resp = requests.get(url, timeout=30,
                            headers={"User-Agent": f"LitRadar/3.1 (mailto:{email})"})
        if resp.status_code == 200:
            for item in resp.json().get("collection", []):
                title = item.get("title", ""); abstract = item.get("abstract", "")
                if not _keyword_match(query, (title + " " + abstract).lower()):
                    continue
                pub_date = item.get("date", "")
                if date_from and pub_date < date_from: continue
                doi = item.get("doi", ""); au = item.get("authors", "").split("; ")
                papers.append({
                    "id": f"doi:{doi}" if doi else f"{server}:{hashlib.md5(title.encode()).hexdigest()[:12]}",
                    "Title": title, "First Author": au[0] if au else "",
                    "All Authors": "; ".join(au), "Journal": f"{server} (preprint)",
                    "Year": pub_date[:4] if pub_date else "", "Pub Date": pub_date,
                    "Abstract": abstract, "MeSH Terms": "", "Keywords": "",
                    "DOI": doi, "PMID": "", "PMCID": "", "Is OA": "Yes",
                    "Source": server, "Citations": 0, "Query Tag": "", "Relevance Score": 0.0,
                })
    except Exception as e:
        logger.error(f"  [{server}] API: {e}")

    logger.info(f"  [{server}] (API) → {len(papers)} 篇")
    return papers


def _extract_keywords_from_query(query: str) -> list[str]:
    """从布尔查询中提取纯关键词，供 paperscraper 使用。"""
    terms = re.findall(r'"([^"]*)"|\b(\w+)\b', query.upper())
    keywords = [t[0] if t[0] else t[1] for t in terms]
    return [k for k in keywords if k not in {"AND", "OR", "NOT"} and len(k) > 2]


def _keyword_match(query: str, text: str) -> bool:
    terms = re.findall(r'"([^"]*)"|\b(\w+)\b', query.upper())
    keywords = [t[0] if t[0] else t[1] for t in terms]
    keywords = [k for k in keywords if k not in {"AND", "OR", "NOT"} and len(k) > 2]
    if not keywords: return True
    matched = sum(1 for k in keywords if k in text.upper())
    return matched >= len(keywords) / 3


# -- 被引次数 --

def enrich_citations(papers: list[dict], config: dict) -> list[dict]:
    dois = [p["DOI"] for p in papers if p.get("DOI")]
    if not dois: return papers
    logger.info(f"  [Semantic Scholar] 查询 {len(dois)} 篇引用...")
    cache = {}; email = config.get("ncbi_email", "")
    for i in range(0, len(dois), 20):
        batch = dois[i:i + 20]
        try:
            resp = requests.post(
                "https://api.semanticscholar.org/graph/v1/paper/batch",
                json={"ids": [f"DOI:{d}" for d in batch]},
                headers={"User-Agent": f"LitRadar/3.1 (mailto:{email})"},
                timeout=60,
            )
            if resp.status_code == 429:
                time.sleep(5)
                resp = requests.post(
                    "https://api.semanticscholar.org/graph/v1/paper/batch",
                    json={"ids": [f"DOI:{d}" for d in batch]},
                    headers={"User-Agent": f"LitRadar/3.1 (mailto:{email})"},
                    timeout=60,
                )
            if resp.status_code == 200:
                for item in resp.json():
                    if item and item.get("externalIds"):
                        d = item["externalIds"].get("DOI", "").lower()
                        if d: cache[d] = item.get("citationCount", 0) or 0
            time.sleep(1)
        except Exception as e:
            logger.debug(f"  [Semantic Scholar] 批次失败: {e}")
    for p in papers:
        d = (p.get("DOI") or "").lower()
        if d in cache: p["Citations"] = cache[d]
    hit = sum(1 for p in papers if p["Citations"] > 0)
    logger.info(f"  [Semantic Scholar] → {hit} 篇有引用数据")
    return papers


# ===========================================================================
# 去重 & 评分
# ===========================================================================


def deduplicate(papers: list[dict]) -> list[dict]:
    """DOI > PMID > 标题模糊匹配 (>0.9)。"""
    seen_doi, seen_pmid = set(), set()
    kept_titles: list[str] = []
    kept = []
    for p in papers:
        doi = (p.get("DOI") or "").strip().lower()
        pmid = (p.get("PMID") or "").strip()
        title = (p.get("Title") or "").strip()

        if doi and doi in seen_doi: continue
        if pmid and pmid in seen_pmid: continue

        # 标题模糊去重
        title_norm = re.sub(r"[^\w\s]", "", title).lower().strip()
        is_dup = False
        for t in kept_titles:
            if difflib.SequenceMatcher(None, title_norm, t).ratio() > 0.9:
                is_dup = True; break
        if is_dup: continue

        if doi: seen_doi.add(doi)
        if pmid: seen_pmid.add(pmid)
        kept_titles.append(title_norm)
        kept.append(p)
    logger.info(f"  去重: {len(papers)} → {len(kept)}")
    return kept


def score_papers(papers: list[dict], keywords: list[str]) -> list[dict]:
    hi_journals = {
        "nature", "science", "cell", "lancet", "nejm", "jama",
        "nature genetics", "nature medicine", "neuron", "bmj",
    }
    for p in papers:
        s = 0.0
        tl = (p.get("Title") or "").lower()
        al = (p.get("Abstract") or "").lower()
        mesh = (p.get("MeSH Terms") or "").lower()
        s += min(sum(2 for k in keywords if k.lower() in tl), 6)
        s += min(sum(1 for k in keywords if k.lower() in al), 3)
        s += min(sum(0.5 for k in keywords if k.lower() in mesh), 1)
        if any(j in (p.get("Journal") or "").lower() for j in hi_journals):
            s += 1.0
        try:
            if int(p.get("Year", "0") or "0") >= 2022: s += 0.5
        except ValueError: pass
        if (p.get("Citations") or 0) > 50: s += 0.5
        p["Relevance Score"] = round(s, 1)
    papers.sort(key=lambda x: x["Relevance Score"], reverse=True)
    return papers


# ===========================================================================
# Excel 导出
# ===========================================================================

COLUMN_ORDER = [
    "#", "Title", "First Author", "All Authors", "Journal", "Year", "Pub Date",
    "Abstract", "MeSH Terms", "Keywords", "DOI", "PMID", "PMCID", "Is OA",
    "Source", "Citations", "Query Tag", "Relevance Score",
    "★ Relevance", "📝 Notes", "📥 PDF Status", "PDF Path",
]


def _write_excel_to_path(papers: list[dict], path: str, config: dict):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    df = pd.DataFrame(papers)
    for col in COLUMN_ORDER:
        if col != "#" and col not in df.columns:
            df[col] = ""
    if "#" in df.columns: df["#"] = range(1, len(df) + 1)
    else: df.insert(0, "#", range(1, len(df) + 1))
    if "Relevance Score" in df.columns:
        df["Relevance Score"] = pd.to_numeric(df["Relevance Score"], errors="coerce").fillna(0)
    if "Citations" in df.columns:
        df["Citations"] = pd.to_numeric(df["Citations"], errors="coerce").fillna(0).astype(int)
    display_cols = [c for c in COLUMN_ORDER if c in df.columns]
    df = df[display_cols]

    wb = openpyxl.Workbook()
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"),
    )
    wrap = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="top")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    oa_font = Font(color="006100", bold=True)

    col_widths = {
        "#": 5, "Title": 50, "First Author": 16, "All Authors": 30,
        "Journal": 25, "Year": 7, "Pub Date": 12, "Abstract": 60,
        "MeSH Terms": 25, "Keywords": 20, "DOI": 30, "PMID": 14,
        "PMCID": 14, "Is OA": 8, "Source": 14, "Citations": 10,
        "Query Tag": 18, "Relevance Score": 16, "★ Relevance": 14,
        "📝 Notes": 20, "📥 PDF Status": 14, "PDF Path": 30,
    }

    def write_sheet(ws, data_df: pd.DataFrame):
        for ci, cn in enumerate(data_df.columns, 1):
            c = ws.cell(row=1, column=ci, value=cn)
            c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, border
        for ri, (_, row) in enumerate(data_df.iterrows()):
            er = ri + 2
            for ci, cn in enumerate(data_df.columns, 1):
                val = row[cn]
                if isinstance(val, float) and pd.isna(val): val = ""
                elif val is None: val = ""
                c = ws.cell(row=er, column=ci, value=val); c.border = border
                if cn == "Abstract": c.alignment = wrap
                elif cn in ("#", "Year", "Citations", "Relevance Score", "Is OA"):
                    c.alignment = center
                else: c.alignment = Alignment(vertical="top")
                if cn == "DOI" and val:
                    try:
                        c.hyperlink = f"https://doi.org/{val}"
                        c.font = Font(color="0563C1", underline="single")
                    except Exception: pass
                if cn == "Is OA" and val == "Yes": c.font = oa_font
        ws.freeze_panes = "A2"
        if len(data_df) > 0: ws.auto_filter.ref = ws.dimensions
        for ci, cn in enumerate(data_df.columns, 1):
            ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(cn, 15)
        if len(data_df) == 0: return
        sc_letter = None
        for ci, cn in enumerate(data_df.columns, 1):
            if cn == "Relevance Score": sc_letter = get_column_letter(ci); break
        if sc_letter:
            lr = len(data_df) + 1; lc = get_column_letter(len(data_df.columns))
            rng = f"A2:{lc}{lr}"
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f"${sc_letter}2>=8"], fill=green_fill))
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f"AND(${sc_letter}2>=5,${sc_letter}2<8)"], fill=yellow_fill))

    # Sheet 1
    ws1 = wb.active; ws1.title = "All Results"; write_sheet(ws1, df)

    # Sheet 2
    high = df[df["Relevance Score"] >= 7].copy()
    if not high.empty: high["#"] = range(1, len(high) + 1)
    ws2 = wb.create_sheet("High Relevance"); write_sheet(ws2, high)

    # Sheet 3
    oa = df[df["Is OA"] == "Yes"].copy()
    if not oa.empty: oa["#"] = range(1, len(oa) + 1)
    ws3 = wb.create_sheet("OA Only"); write_sheet(ws3, oa)

    # Sheet 4
    ws4 = wb.create_sheet("Query Stats")
    stats_headers = ["Query Name", "Query", "Total", "High (≥7)", "OA"]
    for ci, h in enumerate(stats_headers, 1):
        c = ws4.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment = hfont, hfill, halign
    for ri, q in enumerate(config.get("queries", [])):
        nm = q.get("name", ""); qs = q.get("query", "")
        sub = df[df["Query Tag"] == nm]
        vals = [nm, qs, len(sub), len(sub[sub["Relevance Score"] >= 7]),
                len(sub[sub["Is OA"] == "Yes"])]
        for ci, v in enumerate(vals, 1):
            ws4.cell(row=ri + 2, column=ci, value=v)
    sr = len(config.get("queries", [])) + 3
    ws4.cell(row=sr, column=1, value="TOTAL (去重后)").font = Font(bold=True)
    ws4.cell(row=sr, column=3, value=len(df))
    ws4.cell(row=sr, column=4, value=len(high))
    ws4.cell(row=sr, column=5, value=len(oa))
    for i, w in enumerate([22, 55, 10, 14, 10], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    logger.info(f"\n✅ Excel: {path}")
    logger.info(f"   All Results: {len(df)} | High (≥7): {len(high)} | OA: {len(oa)}")


def export_excel(papers: list[dict], output_dir: str, config: dict) -> str:
    path = os.path.join(output_dir, f"results_{datetime.now().strftime('%Y%m%d')}.xlsx")
    _write_excel_to_path(papers, path, config)
    return path


def export_excel_to_path(papers: list[dict], path: str, config: dict) -> str:
    _write_excel_to_path(papers, path, config)
    return path


def export_bibtex(papers: list[dict], output_path: str):
    """导出为 BibTeX 格式。"""
    entries = []
    for p in papers:
        authors = p.get("All Authors", "") or p.get("First Author", "")
        cite_key = (
            f"{p.get('First Author', 'Anon').split()[-1] if p.get('First Author') else 'Anon'}"
            f"{p.get('Year', '0000')[:4]}"
        )
        entry = (
            f"@article{{{cite_key},\n"
            f"  title = {{{p.get('Title', '')}}},\n"
            f"  author = {{{authors.replace('; ', ' and ')}}},\n"
            f"  journal = {{{p.get('Journal', '')}}},\n"
            f"  year = {{{p.get('Year', '')}}},\n"
            f"  doi = {{{p.get('DOI', '')}}},\n"
            f"  pmid = {{{p.get('PMID', '')}}},\n"
            f"  abstract = {{{(p.get('Abstract', '') or '')[:500]}}},\n"
            f"}}"
        )
        entries.append(entry)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n\n".join(entries))
    logger.info(f"✅ BibTeX: {output_path} ({len(entries)} entries)")


# ===========================================================================
# PDF 按需下载
# ===========================================================================

def download_from_excel(excel_path: str, config: dict):
    import openpyxl
    if not os.path.exists(excel_path):
        logger.error(f"文件不存在: {excel_path}"); return

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["All Results"]
    headers = {}
    for ci in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=ci).value
        if v: headers[v] = ci

    doi_c = headers.get("DOI"); title_c = headers.get("Title")
    pmid_c = headers.get("PMID"); pmcid_c = headers.get("PMCID")
    st_c = headers.get("📥 PDF Status"); path_c = headers.get("PDF Path")
    if not st_c: logger.error("找不到 '📥 PDF Status' 列"); return

    targets = []
    for ri in range(2, ws.max_row + 1):
        s = (ws.cell(row=ri, column=st_c).value or "").strip().lower()
        if s in ("need download", "need download"): targets.append(ri)

    if not targets: logger.info("没有标记为 'Need Download' 的条目"); return

    logger.info(f"准备下载 {len(targets)} 篇...")
    email = config.get("unpaywall_email", "") or config.get("ncbi_email", "")
    pdf_dir = config.get("storage", {}).get("pdf_dir", "data/pdfs")
    os.makedirs(pdf_dir, exist_ok=True)

    ok = 0
    for ri in targets:
        doi = ws.cell(row=ri, column=doi_c).value if doi_c else ""
        pmid = ws.cell(row=ri, column=pmid_c).value if pmid_c else ""
        pmcid = ws.cell(row=ri, column=pmcid_c).value if pmcid_c else ""
        title = ws.cell(row=ri, column=title_c).value if title_c else "paper"
        logger.info(f"  [{ri-1}] {str(title)[:80]}...")
        pdf_path = _download_pdf(doi, pmid, pmcid, pdf_dir, str(title), email)
        if pdf_path:
            ws.cell(row=ri, column=st_c, value="Downloaded")
            if path_c: ws.cell(row=ri, column=path_c, value=pdf_path)
            ok += 1; logger.info(f"    ✅ {pdf_path}")
        else:
            ws.cell(row=ri, column=st_c, value="Failed")
            logger.info(f"    ❌ 失败"); time.sleep(1)

    wb.save(excel_path)
    logger.info(f"下载完成: {ok}/{len(targets)}")


def _download_pdf(doi: str, pmid: str, pmcid: str, pdf_dir: str,
                  title: str, email: str) -> Optional[str]:
    safe = re.sub(r"[^\w\s-]", "", str(title))[:60].strip() or "paper"
    if pmcid:
        try:
            resp = requests.get(
                "https://www.ncbi.nlm.nih.gov/pmc/utils/oa/oa.fcgi",
                params={"id": pmcid, "format": "json"}, timeout=30)
            for rec in resp.json().get("records", []):
                for link in rec.get("links", []):
                    if link.get("format") == "pdf" and link.get("href"):
                        data = _fetch_url(link["href"])
                        if data:
                            p = os.path.join(pdf_dir, f"{pmcid}_{safe}.pdf")
                            with open(p, "wb") as f: f.write(data)
                            return p
        except Exception: pass
    if doi:
        try:
            resp = requests.get(
                f"https://api.unpaywall.org/v2/{doi}",
                params={"email": email} if email else {}, timeout=30)
            if resp.status_code == 200:
                url = resp.json().get("best_oa_location", {}).get("url_for_pdf", "")
                if url:
                    data = _fetch_url(url)
                    if data:
                        p = os.path.join(pdf_dir, f"{doi.replace('/', '_')}_{safe}.pdf")
                        with open(p, "wb") as f: f.write(data)
                        return p
        except Exception: pass
    try:
        from paperscraper.utils import get_pdf
        if doi:
            data = get_pdf(doi)
            if data:
                p = os.path.join(pdf_dir, f"{doi.replace('/', '_')}_{safe}.pdf")
                with open(p, "wb") as f: f.write(data)
                return p
    except Exception: pass
    return None


def _fetch_url(url: str, retries: int = 3) -> Optional[bytes]:
    for i in range(retries):
        try:
            resp = requests.get(url, timeout=60, headers={"User-Agent": "LitRadar/3.1"})
            if resp.status_code == 200: return resp.content
            if resp.status_code in (403, 404): return None
        except Exception: pass
        time.sleep(2 ** i)
    return None


# ===========================================================================
# 增量更新
# ===========================================================================

def update_excel(excel_path: str, new_papers: list[dict], config: dict) -> str:
    import openpyxl
    if not os.path.exists(excel_path):
        logger.info("Excel 不存在，创建新文件")
        return export_excel(new_papers, config.get("storage", {}).get("output_dir", "output"), config)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["All Results"]
    headers = {}
    for ci in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=ci).value
        if v: headers[v] = ci

    doi_c = headers.get("DOI")
    existing_dois = set()
    if doi_c:
        for ri in range(2, ws.max_row + 1):
            v = ws.cell(row=ri, column=doi_c).value
            if v: existing_dois.add(str(v).strip().lower())

    truly_new = [p for p in new_papers
                 if (p.get("DOI") or "").strip().lower() not in existing_dois]
    if not truly_new:
        logger.info(f"无新文献（{len(new_papers)} 篇均已存在）")
        return excel_path

    logger.info(f"追加 {len(truly_new)} 篇（{len(new_papers) - len(truly_new)} 篇已存在）")
    base = ws.max_row - 1
    for i, p in enumerate(truly_new):
        ri = ws.max_row + 1
        for ci, cn in enumerate(COLUMN_ORDER, 1):
            val = base + i + 1 if cn == "#" else p.get(cn, "")
            if isinstance(val, float) and pd.isna(val): val = ""
            elif val is None: val = ""
            ws.cell(row=ri, column=ci, value=val)

    wb.save(excel_path)
    logger.info(f"✅ 已更新: {excel_path} (+{len(truly_new)})")
    return excel_path


# ===========================================================================
# 通知系统（完整重写）
# ===========================================================================


class Notifier:
    def __init__(self, config: dict):
        self.cfg = config.get("notifications", {})
        self.desktop = self.cfg.get("desktop", {}).get("enabled", True)
        self.ec = self.cfg.get("email", {})

    def notify_desktop(self, title: str, msg: str):
        if self.desktop:
            try:
                from plyer import notification
                notification.notify(title=title, message=msg[:256], timeout=10)
            except Exception: pass

    def email_digest(self, new_count: int, papers: list[dict],
                     downloaded: int, excel_path: str = None):
        """发送增强版 HTML+纯文本邮件，可选 Excel 附件。"""
        ec = self.ec
        if not ec.get("enabled"): return

        # 无新文献时是否发送
        if new_count == 0 and not ec.get("send_even_if_no_new", False):
            logger.info("无新文献，跳过邮件发送")
            return

        # 密码优先从环境变量读取
        password = ec.get("password") or os.environ.get("LITRADAR_EMAIL_PASS", "")
        if not password:
            logger.error("邮件密码未配置，请设置 config.yaml 或环境变量 LITRADAR_EMAIL_PASS")
            return

        ds = datetime.now().strftime("%Y-%m-%d")
        min_score = ec.get("min_score_highlight", 7)
        max_papers = ec.get("max_papers_in_email", 50)

        # 统计数据
        high_rel = [p for p in papers if p["Relevance Score"] >= min_score]
        oa_count = sum(1 for p in papers if p.get("Is OA") == "Yes")

        # 构建 HTML
        html = self._build_digest_html(new_count, papers, high_rel, oa_count,
                                       downloaded, ds, min_score, max_papers)

        # 构建纯文本 fallback
        plain = self._build_digest_plain(new_count, papers, high_rel, oa_count, ds)

        try:
            # 构建 multipart/mixed 邮件
            msg = MIMEMultipart("mixed")
            msg["From"] = ec.get("username", "")
            msg["To"] = ec.get("to", "")
            msg["Subject"] = (
                f"[LitRadar] {ds} | 新增 {new_count} 篇"
                + (f" | 高相关 {len(high_rel)} 篇" if high_rel else "")
            )

            # HTML + 纯文本双 part
            alt = MIMEMultipart("alternative")
            alt.attach(MIMEText(plain, "plain", "utf-8"))
            alt.attach(MIMEText(html, "html", "utf-8"))
            msg.attach(alt)

            # Excel 附件
            if excel_path and ec.get("attach_excel", True) and os.path.exists(excel_path):
                with open(excel_path, "rb") as f:
                    att = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    att.set_payload(f.read())
                encoders.encode_base64(att)
                att.add_header(
                    "Content-Disposition",
                    f'attachment; filename="litradar_{ds}.xlsx"'
                )
                msg.attach(att)

            # 发送
            s = smtplib.SMTP(ec.get("smtp_server", ""), ec.get("smtp_port", 587))
            s.starttls()
            s.login(ec.get("username", ""), password)
            s.send_message(msg)
            s.quit()
            logger.info(f"邮件已发送至 {ec.get('to', '')}")
        except Exception as e:
            logger.error(f"邮件发送失败: {e}")

    def _build_digest_html(self, new_count, papers, high_rel, oa_count,
                           downloaded, ds, min_score, max_papers):
        # 统计 Banner
        banner = f"""
        <div style="background:#2F5496;color:#fff;padding:20px;border-radius:8px;margin-bottom:20px;">
            <h2 style="margin:0 0 10px 0;">📚 LitRadar 文献日报 — {ds}</h2>
            <table style="color:#fff;font-size:14px;">
                <tr>
                    <td style="padding:4px 20px 4px 0;">📄 新增文献</td><td><strong>{new_count}</strong> 篇</td>
                    <td style="padding:4px 20px 4px 40px;">🟢 开放获取</td><td><strong>{oa_count}</strong> 篇</td>
                    <td style="padding:4px 20px 4px 40px;">⭐ 高相关</td><td><strong>{len(high_rel)}</strong> 篇</td>
                    <td style="padding:4px 20px 4px 40px;">📥 已下载</td><td><strong>{downloaded}</strong> 篇</td>
                </tr>
            </table>
        </div>
        """

        # 高相关文献区
        high_section = ""
        if high_rel:
            high_section = (
                f'<h3 style="color:#2F5496;">⭐ 高相关文献（Score ≥ {min_score}）</h3>'
                f'<table style="width:100%;border-collapse:collapse;">'
            )
            for p in high_rel[:max_papers]:
                title = p.get("Title", "")
                authors = p.get("First Author", "")
                journal = p.get("Journal", "")
                year = p.get("Year", "")
                score = p.get("Relevance Score", 0)
                abstract = (p.get("Abstract") or "")[:150]
                doi = p.get("DOI", "")
                oa = "🟢" if p.get("Is OA") == "Yes" else "🔒"

                high_section += f"""
                <tr>
                    <td style="padding:12px;border-bottom:1px solid #eee;vertical-align:top;">
                        <div style="font-size:14px;font-weight:bold;margin-bottom:4px;">
                            {oa} {title}
                        </div>
                        <div style="font-size:12px;color:#666;margin-bottom:4px;">
                            {authors} · {journal} · {year} · Score: {score}
                        </div>
                        <div style="font-size:12px;color:#444;margin-bottom:4px;">
                            {abstract}{'...' if len(p.get('Abstract', '') or '') > 150 else ''}
                        </div>
                        {f'<a href="https://doi.org/{doi}" style="font-size:12px;">🔗 {doi}</a>' if doi else ''}
                    </td>
                </tr>
                """
            high_section += "</table>"

        # 其余文献折叠展示
        rest = [p for p in papers if p not in high_rel]
        rest_section = ""
        if rest:
            rest_section = (
                f'<h3 style="color:#666;">📋 其余新增文献（Score < {min_score}）</h3>'
                f'<table style="width:100%;border-collapse:collapse;">'
            )
            for p in rest[:max_papers]:
                title = p.get("Title", "")
                abstract = (p.get("Abstract") or "")[:80]
                doi = p.get("DOI", "")
                rest_section += f"""
                <tr>
                    <td style="padding:6px;border-bottom:1px solid #f0f0f0;">
                        <div style="font-size:13px;">{title}</div>
                        <div style="font-size:11px;color:#888;">{abstract}...</div>
                        {f'<a href="https://doi.org/{doi}" style="font-size:11px;">{doi}</a>' if doi else ''}
                    </td>
                </tr>
                """
            rest_section += "</table>"

        return f"""<html><body style="font-family:Arial,sans-serif;max-width:750px;margin:0 auto;padding:10px;">
{banner}{high_section}{rest_section}
<hr><p style="color:#999;font-size:11px;">LitRadar v3.1 自动生成 | 配置: config.yaml</p>
</body></html>"""

    def _build_digest_plain(self, new_count, papers, high_rel, oa_count, ds):
        lines = [
            f"LitRadar 文献日报 — {ds}",
            f"─────────────────────────────",
            f"新增: {new_count} 篇 | OA: {oa_count} 篇 | 高相关: {len(high_rel)} 篇",
            f"",
        ]
        if high_rel:
            lines.append("⭐ 高相关文献:")
            for p in high_rel[:20]:
                lines.append(f"  {p.get('Title','')[:100]}")
                lines.append(f"  {p.get('First Author','')} | {p.get('Journal','')} | Score: {p.get('Relevance Score',0)}")
                if p.get("DOI"): lines.append(f"  https://doi.org/{p.get('DOI')}")
                lines.append("")
        return "\n".join(lines)


# ===========================================================================
# 检索流水线
# ===========================================================================


def run_search(config: dict, query_name: str = None) -> list[dict]:
    queries = config.get("queries", [])
    date_from = config.get("date_from", "2020-01-01")
    if query_name:
        queries = [q for q in queries if q.get("name") == query_name]
        if not queries:
            logger.error(f"未找到关键词组: {query_name}"); sys.exit(1)

    all_papers = []
    for qc in queries:
        nm = qc.get("name", "unnamed"); query = qc.get("query", "")
        sources = qc.get("sources", ["pubmed"])
        logger.info(f"\n{'='*50}\n📌 [{nm}] {query}\n   数据源: {sources}\n{'='*50}")

        group = []
        fetchers = {
            "pubmed": (fetch_pubmed, query, config, date_from),
            "europe_pmc": (fetch_europe_pmc, query, config, date_from),
            "biorxiv": (fetch_biorxiv, query, config, date_from),
            "medrxiv": (fetch_medrxiv, query, config, date_from),
        }
        with ThreadPoolExecutor(max_workers=4) as ex:
            futs = {src: ex.submit(f, *args) for src, (f, *args) in fetchers.items() if src in sources}
            for src, fut in futs.items():
                try: group.extend(fut.result())
                except Exception as e: logger.error(f"  [{src}] 异常: {e}")

        for p in group: p["Query Tag"] = nm
        logger.info(f"  组合并: {len(group)} 篇")
        all_papers.extend(group)

    all_papers = deduplicate(all_papers)
    sk = config.get("score_keywords", [])
    if sk:
        logger.info(f"📊 计算相关性评分 ({len(sk)} 个核心词)...")
        all_papers = score_papers(all_papers, sk)
    if config.get("fetch_citations"):
        all_papers = enrich_citations(all_papers, config)
        if sk: all_papers = score_papers(all_papers, sk)
    return all_papers


# ===========================================================================
# 追踪调度
# ===========================================================================


class Tracker:
    def __init__(self, config_path: str):
        self.config_path = config_path
        self.config = load_config(config_path)
        sc = self.config.get("storage", {})
        self.storage = Storage(sc.get("db_path", "data/litradar.db"))
        self.notifier = Notifier(self.config)
        self._scheduler = None

    def run_once(self) -> dict:
        queries = self.config.get("queries", [])
        stats = {"queries": 0, "found": 0, "new": 0, "downloaded": 0}
        all_new, all_papers = [], []

        for qc in queries:
            nm = qc.get("name", "unnamed"); query = qc.get("query", "")
            sources = qc.get("sources", ["pubmed"])
            last = self.storage.get_last_run(nm)
            date_from = ""
            if last:
                try:
                    dt = datetime.fromisoformat(last) - timedelta(days=1)
                    date_from = dt.strftime("%Y-%m-%d")
                except Exception: pass

            logger.info(f"\n--- [{nm}] {query} ---")
            group = []
            fetchers = {
                "pubmed": (fetch_pubmed, query, self.config, date_from),
                "europe_pmc": (fetch_europe_pmc, query, self.config, date_from),
                "biorxiv": (fetch_biorxiv, query, self.config, date_from),
                "medrxiv": (fetch_medrxiv, query, self.config, date_from),
            }
            with ThreadPoolExecutor(max_workers=4) as ex:
                futs = {src: ex.submit(f, *args) for src, (f, *args) in fetchers.items() if src in sources}
                for src, fut in futs.items():
                    try: group.extend(fut.result())
                    except Exception as e: logger.error(f"  [{src}]: {e}")

            for p in group: p["Query Tag"] = nm
            group = deduplicate(group)

            new_count = 0
            for p in group:
                if not self.storage.exists(p.get("DOI", ""), p.get("PMID", "")):
                    if self.storage.upsert(p):
                        new_count += 1
                        all_new.append(p)
                all_papers.append(p)

            sk = self.config.get("score_keywords", [])
            if sk: all_new = score_papers(all_new, sk)

            self.storage.log_track_run(nm, query, len(group), new_count)
            stats["queries"] += 1; stats["found"] += len(group); stats["new"] += new_count
            logger.info(f"  [{nm}] {len(group)} 篇, 新增 {new_count}")

        # 下载 OA
        if all_new:
            dl_cfg = {"unpaywall_email": self.config.get("unpaywall_email", ""),
                       "ncbi_email": self.config.get("ncbi_email", "")}
            pdf_dir = self.config.get("storage", {}).get("pdf_dir", "data/pdfs")
            for p in [p for p in all_new if p.get("Is OA") == "Yes"]:
                path = _download_pdf(p.get("DOI", ""), p.get("PMID", ""),
                                     p.get("PMCID", ""), pdf_dir,
                                     p.get("Title", ""), dl_cfg["unpaywall_email"])
                if path:
                    self.storage.mark_downloaded(p["id"], path)
                    stats["downloaded"] += 1

        # 通知 + 邮件
        if stats["new"] > 0:
            self.notifier.notify_desktop(
                f"LitRadar — {stats['new']} 篇新文献",
                f"发现 {stats['new']} 篇，下载 {stats['downloaded']} 篇")

            # 导出 Excel 供邮件附件
            od = self.config.get("storage", {}).get("output_dir", "output")
            excel_path = export_excel(all_papers, od, self.config)
            self.notifier.email_digest(
                stats["new"], all_new, stats["downloaded"],
                excel_path=excel_path)

        return stats

    def start_daemon(self):
        try:
            from apscheduler.schedulers.background import BackgroundScheduler
            from apscheduler.triggers.cron import CronTrigger
        except ImportError:
            logger.error("APScheduler 未安装: pip install apscheduler"); return

        sc = self.config.get("scheduler", {})
        rt = sc.get("daily_run_time", "08:00"); tz = sc.get("timezone", "Asia/Shanghai")
        h, m = rt.split(":")
        self._scheduler = BackgroundScheduler()
        self._scheduler.add_job(
            self._job, CronTrigger(hour=int(h), minute=int(m), timezone=tz),
            id="litradar_daily", replace_existing=True)
        logger.info(f"⏰ 定时追踪已启动: 每天 {rt} ({tz})")
        logger.info("   提示: 在服务器上建议用 nohup python litradar.py watch --daemon & 或 systemd service")
        self._scheduler.start()
        try:
            import signal
            signal.signal(signal.SIGINT, lambda *_: (self._scheduler.shutdown(wait=False), sys.exit(0)))
            signal.signal(signal.SIGTERM, lambda *_: (self._scheduler.shutdown(wait=False), sys.exit(0)))
            signal.pause()
        except (KeyboardInterrupt, SystemExit):
            logger.info("守护进程已停止")

    def _job(self):
        logger.info("=== 定时追踪开始 ===")
        try:
            stats = self.run_once()
            logger.info(f"=== 完成: {stats} ===")
        except Exception as e:
            logger.error(f"定时任务异常: {e}")


# ===========================================================================
# CLI
# ===========================================================================


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="LitRadar v3.1 — 融合版文献追踪系统")
    p.add_argument("--config", "-c", default="config.yaml")
    sp = p.add_subparsers(dest="cmd")

    sp1 = sp.add_parser("search", help="一次性检索")
    sp1.add_argument("--query", "-q", required=True)
    sp1.add_argument("--date-from"); sp1.add_argument("--max", type=int, default=200)
    sp1.add_argument("--sources"); sp1.add_argument("--export", action="store_true")
    sp1.add_argument("--send-email", action="store_true", help="检索后发送邮件报告")
    sp1.set_defaults(func=cmd_search)

    sp2 = sp.add_parser("watch", help="文献追踪")
    sp2.add_argument("--daemon", action="store_true")
    sp2.add_argument("--run-now", action="store_true")
    sp2.set_defaults(func=cmd_watch)

    sp3 = sp.add_parser("download", help="按需下载 PDF（从 Excel）")
    sp3.add_argument("--excel", required=True)
    sp3.set_defaults(func=cmd_download)

    sp4 = sp.add_parser("export", help="导出 Excel")
    sp4.add_argument("--output", "-o"); sp4.add_argument("--format", "-f", default="excel")
    sp4.set_defaults(func=cmd_export)

    sp5 = sp.add_parser("stats", help="数据库统计")
    sp5.set_defaults(func=cmd_stats)

    sp6 = sp.add_parser("serve", help="启动 REST API 服务供鸿蒙客户端调用")
    sp6.add_argument("--host", default="127.0.0.1")
    sp6.add_argument("--port", type=int, default=7749)
    sp6.set_defaults(func=lambda args: serve_api(args.config, args.host, args.port))

    return p


def cmd_search(args):
    config = load_config(args.config)
    sc = config.get("storage", {})
    storage = Storage(sc.get("db_path", "data/litradar.db"))
    if args.date_from: config["date_from"] = args.date_from
    if args.max: config["max_results_per_source"] = args.max
    sources = args.sources.split(",") if args.sources else None
    config["queries"] = [{"name": "cli_search", "query": args.query,
                          "sources": sources or ["pubmed", "europe_pmc", "biorxiv"]}]
    papers = run_search(config)
    saved = sum(1 for p in papers if storage.upsert(p))
    logger.info(f"入库 {saved} 篇")
    for i, p in enumerate(papers[:20], 1):
        oa = "🟢" if p.get("Is OA") == "Yes" else "🔒"
        logger.info(f"{i:3d}. [{oa}] {p.get('Title','')[:100]}")
    if len(papers) > 20: logger.info(f"... 还有 {len(papers)-20} 篇")

    path = None
    if args.export:
        path = export_excel(papers, sc.get("output_dir", "output"), config)

    if args.send_email:
        notifier = Notifier(config)
        notifier.email_digest(new_count=saved, papers=papers, downloaded=0,
                              excel_path=path)


def cmd_watch(args):
    tracker = Tracker(args.config)
    if args.run_now:
        logger.info("手动触发追踪...")
        stats = tracker.run_once()
        logger.info(f"完成: {stats}")
    elif args.daemon:
        tracker.start_daemon()
    else:
        logger.info("请指定 --daemon 或 --run-now")


def cmd_download(args):
    config = load_config(args.config)
    download_from_excel(args.excel, config)


def cmd_export(args):
    config = load_config(args.config)
    sc = config.get("storage", {})
    storage = Storage(sc.get("db_path", "data/litradar.db"))
    papers = storage.get_all()
    if not papers:
        logger.info("数据库为空，请先执行 search 或 watch --run-now"); return

    fmt = getattr(args, "format", "excel")
    od = getattr(args, "output", None) or sc.get("output_dir", "output")

    if fmt == "bibtex":
        export_bibtex(papers, od if od.endswith(".bib") else od + ".bib")
    else:
        if od.endswith(".xlsx"):
            export_excel_to_path(papers, od, config)
        else:
            path = export_excel(papers, od, config)
            if config.get("export", {}).get("auto_open_excel"):
                try:
                    if sys.platform == "darwin": subprocess.call(["open", path])
                    elif sys.platform == "win32": os.startfile(path)
                    else: subprocess.call(["xdg-open", path])
                except Exception: pass


def cmd_stats(args):
    config = load_config(args.config)
    sc = config.get("storage", {})
    storage = Storage(sc.get("db_path", "data/litradar.db"))
    s = storage.get_stats()
    print(f"""
=== LitRadar v3.1 数据库统计 ===
  论文总数:      {s['total']}
  OA 论文:       {s['oa']}
  已下载 PDF:    {s['downloaded']}
  高相关 (≥7):   {s['high_relevance']}
""")
    history = storage.get_track_history()
    if history:
        print("最近追踪记录:")
        for h in history[:10]:
            print(f"  {h['run_at'][:19]} | {h['query_name']:25s} | 发现 {h['papers_found']:3d} | 新增 {h['papers_new']:3d}")


# ─── REST API Server ───────────────────────────────────────────
def serve_api(config_path: str = "config.yaml", host: str = "127.0.0.1", port: int = 7749):
    """
    启动 Flask REST 服务，供鸿蒙客户端调用。
    端口固定为 7749，不可通过参数修改。
    """
    from flask import Flask, jsonify, request as freq
    cfg = load_config(config_path)
    storage = Storage(cfg.get("storage", {}).get("db_path", "data/litradar.db"))

    app = Flask("litradar_api")

    @app.route("/api/papers", methods=["GET"])
    def api_papers():
        query_name = freq.args.get("query_name", None)   # 可选过滤
        limit      = int(freq.args.get("limit", 200))
        with storage._conn() as c:
            if query_name:
                rows = c.execute(
                    "SELECT * FROM papers WHERE [Query Tag]=? ORDER BY [Relevance Score] DESC LIMIT ?",
                    (query_name, limit)
                ).fetchall()
            else:
                rows = c.execute(
                    "SELECT * FROM papers ORDER BY [Relevance Score] DESC, [Pub Date] DESC LIMIT ?",
                    (limit,)
                ).fetchall()
        return jsonify([dict(r) for r in rows])

    @app.route("/api/queries", methods=["GET"])
    def api_queries():
        """返回 config.yaml 中配置的所有 query 名称和 query 字符串"""
        queries = cfg.get("queries", [])
        return jsonify([{"name": q.get("name",""), "query": q.get("query","")} for q in queries])

    @app.route("/api/stats", methods=["GET"])
    def api_stats():
        return jsonify(storage.get_stats())

    @app.route("/api/run", methods=["POST"])
    def api_run():
        """
        触发一次 run_search。
        POST body JSON: {"query_name": "AD_APOE"}  ← 可选，不传则跑所有
        注意：此接口是同步阻塞的，鸿蒙端需设置超时 >= 120 秒。
        """
        body = freq.get_json(silent=True) or {}
        query_name = body.get("query_name", None)
        papers = run_search(cfg, query_name)
        saved = sum(1 for p in papers if storage.upsert(p))
        return jsonify({"found": len(papers), "saved": saved})

    @app.route("/api/health", methods=["GET"])
    def api_health():
        return jsonify({"status": "ok", "version": "3.1"})

    app.run(host=host, port=port, debug=False, threaded=True)


def main():
    parser = build_parser()
    args = parser.parse_args()
    if not args.cmd: parser.print_help(); return
    os.makedirs("logs", exist_ok=True); os.makedirs("data/pdfs", exist_ok=True)
    try:
        from loguru import logger as _l
        _l.add("logs/litradar.log", rotation="10 MB", retention="30 days",
               format="{time:YYYY-MM-DD HH:mm:ss} | {level} | {message}")
    except Exception: pass
    args.func(args)


if __name__ == "__main__":
    main()
